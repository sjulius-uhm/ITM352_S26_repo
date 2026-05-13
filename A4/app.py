
#Company Financial Data Web Scraper and Analysis Tool

#This Flask application lets users log in, analyze company financial data, compare multiple companies, generate charts, and export results to CSV/Excel.

#Libraries Used
from flask import Flask, render_template, request, send_file, redirect, url_for, session

#"For basic sentiment analysis of company news headlines. 
#Got this idea from A.I using prompt 
# "Make a sentimental analysis function that looks at recent news about a company and estimates if the overall sentiment is positive, negative, or neutral."
from textblob import TextBlob
from functools import wraps
import os
import json
from datetime import datetime

# Data processing and visualization libraries
import pandas as pd
import matplotlib
matplotlib.use("Agg")  # Lets matplotlib save charts without opening a window
import matplotlib.pyplot as plt
import numpy as np

# library import for Yahoo Finance API without actually relying on API 
try:
    import yfinance as yf
except ImportError:
    yf = None

#Website set up and configuration
app = Flask(__name__)
app.secret_key = "fin-analytix-secret-key-2026"

#File paths and folders for folders for user data, outputs, and charts.
USERS_FILE = "users.json"
OUTPUT_FOLDER = "outputs"
CHART_FOLDER = os.path.join("static", "charts")
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
os.makedirs(CHART_FOLDER, exist_ok=True)


# Rank folders for organizing saved analyses
RANK_FOLDERS = {
    "High Rank": {"label": "High Rank (Score 80-100)", "min": 80, "max": 100},
    "Stable": {"label": "Stable (Score 50-79)", "min": 50, "max": 79},    
    "Risky": {"label": "Risky (Score 0-49)", "min": 0, "max": 49},
}


# Create rank subfolders inside outputs
for folder_name in RANK_FOLDERS:
    os.makedirs(os.path.join(OUTPUT_FOLDER, folder_name), exist_ok=True)


# File to persist analysis history
HISTORY_FILE = os.path.join(OUTPUT_FOLDER, "analysis_history.json")


# ---- Sentiment Analysis Function ----


def get_sentiment_analysis(ticker_symbol):
    
    #Gets recent company news from yfinance and uses TextBlob to estimate
   # basic sentiment from the article headlines.
    # Returns: (Label, Avg Polarity Score, Good News List, Bad News List)

    #Logic statements to categorize sentiment based on polarity/Either good or bad score
    # Had to use A.I to help with this one, as sentiment analysis was hard to quantify. I used the prompt:
    #"How to quantify sentiment scores into categories like Positive, Negative, and Neutral? in Python"
    try:
        company = yf.Ticker(ticker_symbol)
        news = company.news

        if not news:
            return "No Recent News", 0.0, [], []

        total_polarity = 0
        good_news = []
        bad_news = []
        analyzed_count = 0

#Looks at the top 10 news items to avoid skewing the sentiment score and limit to most recent 10 and extract title
        for item in news[:10]:

            # yfinance news format can vary, so this checks multiple places and s
            title = item.get("title")

            if title is None and "content" in item:
                title = item["content"].get("title")

            if title is None:
                continue

            analysis = TextBlob(title)
            polarity = analysis.sentiment.polarity

            total_polarity += polarity
            analyzed_count += 1
#Polarity is a score from -1 to 1, where -1 is very negative, 0 is neutral, and 1 is very positive.
            if polarity > 0.1:
                good_news.append(title)
            elif polarity < -0.1:
                bad_news.append(title)

        if analyzed_count == 0:
            return "No Headlines Found", 0.0, [], []

        avg_polarity = total_polarity / analyzed_count

        if avg_polarity > 0.05:
            sentiment_label = "Positive"
        elif avg_polarity < -0.05:
            sentiment_label = "Negative"
        else:
            sentiment_label = "Neutral"

        return sentiment_label, round(avg_polarity, 2), good_news, bad_news

    except Exception as error:
        print("Sentiment analysis error:", error)
        return "Not available", 0.0, [], []

#---- User and History Management Functions ----
#Load Data that was saved to disk and save new data to disk. This includes analysis history.
def load_history():
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE, "r") as f:
            return json.load(f)
    return []

#Saves data to disk in a JSON file. This is uses analysis history.
def save_history(history):
    with open(HISTORY_FILE, "w") as f:
        json.dump(history, f, indent=2)

#For user loading
def load_users():
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, "r") as f:
            return json.load(f)
    return {}

#For user saving
def save_users(users):
    with open(USERS_FILE, "w") as f:
        json.dump(users, f, indent=2)


#Safety and Login implementation to protect route requiring login
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "username" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


# ---- Watchlist Helper Function ----


def add_to_watchlist(username, ticker):
    #Adds a ticker to the user's specific watchlist. Had to use A.I to help with with making watchlist and the logical statement
    users = load_users()
    if username in users:
        # Check if user data is a dict (new format) or string (old password-only format)
        if isinstance(users[username], str):
            users[username] = {"password": users[username], "watchlist": []}
        
        if "watchlist" not in users[username]:
            users[username]["watchlist"] = []

        if ticker not in users[username]["watchlist"]:
            users[username]["watchlist"].append(ticker)
            save_users(users)
            return True
    return False


# ---- Authentication and Login Routes ----


#Hanldes user login and creating session tied to username. Also handles user registration and logout.
@app.route("/login", methods=["GET", "POST"])
def login():
    """Login page."""
    if "username" in session:
        return redirect(url_for("home"))
#error handling for login form submission and validation. Also checks user credentials against saved users.
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        if not username or not password:
            error = "Please fill in both fields."
        else:
            users = load_users()

            if username in users:
                user_data = users[username]

                # Support both old and new user formats
                if isinstance(user_data, str):
                    saved_password = user_data
                else:
                    saved_password = user_data.get("password")

                if saved_password == password:
                    session["username"] = username
                    return redirect(url_for("home"))
                else:
                    error = "Invalid username or password."
            else:
                error = "Invalid username or password."

    return render_template("login.html", error=error)

#Handles user registration, including validation for username and password, and checks for existing usernames. Creates a new user entry in the users file if registration is successful.
@app.route("/register", methods=["GET", "POST"])
def register():
    """Registration page."""
    if "username" in session:
        return redirect(url_for("home"))

    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        confirm = request.form.get("confirm", "").strip()

        if not username or not password:
            error = "Please fill in all fields."
        elif len(password) < 4:
            error = "Password must be at least 4 characters."
        elif password != confirm:
            error = "Passwords do not match."
        else:
            users = load_users()
            if username in users:
                error = "That username is already taken."
            else:
                users[username] = {
                    "password": password,
                    "watchlist": []
                }
                save_users(users)
                session["username"] = username
                return redirect(url_for("home"))

    return render_template("register.html", error=error)

#Ending the sesson for user and redirecting to login page.
@app.route("/logout")
def logout():
    """Log the user out."""
    session.pop("username", None)
    return redirect(url_for("login"))


# ---- Formatting ----

#In order to format the currency and numbers to be displayed on webpage results  and to handle missing values and invalid  data without crashing webpage
def safe_value(value):
    #Converts missing or invalid values into None so the app does not crash.
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except TypeError:
        pass
    return value


def format_currency(value):
    #Formats large financial numbers for display on the webpage.
    value = safe_value(value)
    if value is None:
        return "Not available"
    try:
        return "${:,.0f}".format(float(value))
    except (ValueError, TypeError):
        return "Not available"


def format_number(value):
    #Formats ratios and percentages for display.
    value = safe_value(value)
    if value is None:
        return "Not available"
    try:
        return "{:,.2f}".format(float(value))
    except (ValueError, TypeError):
        return "Not available"


def format_percent(value):
    #Formats a decimal ratio as a percentage string."
    value = safe_value(value)
    if value is None:
        return "Not available"
    try:
        return "{:.2f}%".format(float(value) * 100)
    except (ValueError, TypeError):
        return "Not available"


# ---- Data Retrieval Functions ----


def clean_ticker(ticker):
   #Searching for company using ticker as that's what yahoo finance uses for its search engine
    ticker = ticker.strip()
    ticker = ticker.upper()
    return ticker


def get_statement_data(ticker):
    
    #Gets financial statement data from Yahoo Finance using yfinance.
    #Returns balance sheet, income statement, and cash flow statement.
    try:
        company = yf.Ticker(ticker)

        balance_sheet = company.balance_sheet
        income_statement = company.financials
        cash_flow = company.cashflow

        return balance_sheet, income_statement, cash_flow

    except Exception as error:
        print("Error getting financial data for", ticker, ":", error)
        return None, None, None


def clean_statement(statement):
    # Cleans one financial statement.
    # Replaces missing values with 0.

    if statement is None or statement.empty:
        return pd.DataFrame()

    statement = statement.fillna(0)
    return statement


def get_value(statement, row_name):
    # Gets one value from a financial statement by row name.
    # Uses the most recent column of data.
    try:
        value = statement.loc[row_name].iloc[0]
        return value

    except Exception:
        return 0


def build_clean_financial_dict(ticker):

    #Builds a clean financial dictionary for one company.
    #This dictionary can be used for calculations, comparisons, and display.
    ticker = clean_ticker(ticker)

    balance_sheet, income_statement, cash_flow = get_statement_data(ticker)

    if balance_sheet is None or balance_sheet.empty:
        print("Invalid ticker or no data found for:", ticker)
        return None

    balance_sheet = clean_statement(balance_sheet)
    income_statement = clean_statement(income_statement)
    cash_flow = clean_statement(cash_flow)

#Complies all the financial data that is scapped from the financial statements into one dictionary
    financial_data = {
        "ticker": ticker,

        # Balance Sheet
        "total_assets": get_value(balance_sheet, "Total Assets"),
        "total_liabilities": get_value(balance_sheet, "Total Liabilities Net Minority Interest"),
        "stockholders_equity": get_value(balance_sheet, "Stockholders Equity"),
        "current_assets": get_value(balance_sheet, "Current Assets"),
        "current_liabilities": get_value(balance_sheet, "Current Liabilities"),

        # Income Statement
        "total_revenue": get_value(income_statement, "Total Revenue"),
        "net_income": get_value(income_statement, "Net Income"),
        "gross_profit": get_value(income_statement, "Gross Profit"),
        "operating_income": get_value(income_statement, "Operating Income"),

        # Cash Flow
        "operating_cash_flow": get_value(cash_flow, "Operating Cash Flow"),
        "free_cash_flow": get_value(cash_flow, "Free Cash Flow"),
    }

    return financial_data


def get_multiple_companies_data(ticker_list):

#    Gets cleaned financial data for multiple companies.
  # Used for company comparisons.
  # Invalid tickers are skipped but the rest of the valid companies will still be processed and returned.
    
    company_data_list = []

    for ticker in ticker_list:
        ticker = clean_ticker(ticker)

        if ticker == "":
            continue

        print("Getting data for:", ticker)

        financial_data = build_clean_financial_dict(ticker)

        if financial_data is not None:
            company_data_list.append(financial_data)
        else:
            print("Skipping invalid ticker:", ticker)

    return company_data_list


# ---- Data Mapping Function ----


def get_financial_data(ticker_symbol):
    # Creates the main company data dictionary used by the app.
    # This function calls build_clean_financial_dict() to retrieve financial
    # statement values, then adds extra company information from yfinance so
    # the templates have one consistent dictionary to use.
    

    ticker_symbol = clean_ticker(ticker_symbol)
    raw = build_clean_financial_dict(ticker_symbol)

#Error handling for invalid ticker symbols
    if raw is None:
        raise ValueError("No company data was found. Check the ticker symbol and try again.")

    company = yf.Ticker(ticker_symbol)
    info = company.info

#How data will be presented in the webpage results. This is the main dictionary that will be used for display, calculations, and comparisons. It combines the financial statement data with additional company info from yfinance.
    data = {
        "Ticker": ticker_symbol,
        "Company Name": info.get("longName", ticker_symbol),
        "Sector": info.get("sector", "N/A"),
        "Industry": info.get("industry", "N/A"),
        "Current Price": info.get("currentPrice") or info.get("regularMarketPrice"),
        "Market Cap": info.get("marketCap"),
        "Total Revenue": raw.get("total_revenue"),
        "Gross Profits": raw.get("gross_profit"),
        "Net Income": raw.get("net_income"),
        "Total Cash": info.get("totalCash"),
        "Total Debt": raw.get("total_liabilities"),
        "Book Value": info.get("bookValue"),
        "Profit Margin": info.get("profitMargins"),
        "Return on Equity": info.get("returnOnEquity"),
        "Debt to Equity": info.get("debtToEquity"),
        "Trailing PE": info.get("trailingPE"),
        "Forward PE": info.get("forwardPE"),
        "Recommendation": info.get("recommendationKey"),
        "Total Assets": raw.get("total_assets"),
        "Total Liabilities": raw.get("total_liabilities"),
        "Stockholders Equity": raw.get("stockholders_equity"),
        "Current Assets": raw.get("current_assets"),
        "Current Liabilities": raw.get("current_liabilities"),
        "Operating Income": raw.get("operating_income"),
        "Operating Cash Flow": raw.get("operating_cash_flow"),
        "Free Cash Flow": raw.get("free_cash_flow"),
    }

    return data


# ---- Financial Analysis Functions ----


def calculate_ratios(data):
    #Calculate financial ratios from the company data dictionary.
    # Extract values safely in order to do calculations
    net_income = safe_value(data.get("Net Income"))
    revenue = safe_value(data.get("Total Revenue"))
    total_cash = safe_value(data.get("Total Cash"))
    total_debt = safe_value(data.get("Total Debt"))
    market_cap = safe_value(data.get("Market Cap"))

    ratios = {}

    # Net Profit Margin: (Net Income / Total Revenue)
    if net_income is not None and revenue and revenue != 0:
        ratios["Net Profit Margin"] = net_income / revenue
    else:
        ratios["Net Profit Margin"] = None

    # Cash to Debt Ratio: (Total Cash / Total Debt)
    if total_cash is not None and total_debt and total_debt != 0:
        ratios["Cash to Debt Ratio"] = total_cash / total_debt
    else:
        ratios["Cash to Debt Ratio"] = None

    # Market Cap to Revenue: (Market Cap / Total Revenue)
    if market_cap is not None and revenue and revenue != 0:
        ratios["Market Cap to Revenue"] = market_cap / revenue
    else:
        ratios["Market Cap to Revenue"] = None

    # Directly map existing ratios from data
    ratios["Debt to Equity"] = safe_value(data.get("Debt to Equity"))
    ratios["Trailing PE"] = safe_value(data.get("Trailing PE"))
    ratios["Return on Equity"] = safe_value(data.get("Return on Equity"))

    return ratios


def categorize_company(data, ratios):
    
    #Categorize a company based on its data and ratios.Into Losing, Growth/Strong Profit, Possible Value, or Neutral.
    #Scoring system based on financial indicators and each assigned a point valuye then added up to be catogorized.
    #Had to use A.I to help with scoring system and making sure that it adds up to 100 and is balanced across different financial indicators. I used the prompt:"How to create a balanced scoring system that categorizes companies into Losing, Growth/Strong Profit, Possible Value, or Neutral based on financial indicators like Net Income, Profit Margin, and P/E Ratio? The scoring system should add up to 100 points and be weighted appropriately across different indicators."

    net_income = safe_value(data.get("Net Income"))
    profit_margin = safe_value(ratios.get("Net Profit Margin"))
    pe_ratio = safe_value(ratios.get("Trailing PE"))

    if net_income is not None and net_income < 0:
        return "Losing Company"
    
    if profit_margin is not None and profit_margin > 0.15:
        return "Growth / Strong Profit Company"
    
    if pe_ratio is not None and 0 < pe_ratio <= 20:
        return "Possible Value Company"

    return "Neutral / Needs More Review"


def calculate_score(data, ratios):
    #Calculate a financial health score from 0-100.
    
    score = 50  # Start at Neutral
    
    # 1. Profitability (Up to +15 or -15)
    margin = safe_value(ratios.get("Net Profit Margin"))
    if margin is not None:
        if margin > 0.20: score += 15
        elif margin > 0.10: score += 5
        elif margin < 0: score -= 15

    # 2. Return on Equity (Up to +15)
    roe = safe_value(ratios.get("Return on Equity"))
    if roe is not None:
        if roe > 0.15: score += 15
        elif roe > 0.08: score += 5

    # 3. Debt Management (Up to +10 or -20)
    d_e = safe_value(ratios.get("Debt to Equity"))
    if d_e is not None:
        if d_e < 50: score += 10    # Low debt
        elif d_e > 150: score -= 20  # High debt

    # 4. Cash Position (Up to +10)
    cash_debt = safe_value(ratios.get("Cash to Debt Ratio"))
    if cash_debt is not None and cash_debt > 1.0:
        score += 10 # Can pay off all debt with cash

    # 5. Valuation (Up to +10 or -10)
    pe = safe_value(ratios.get("Trailing PE"))
    if pe is not None:
        if 0 < pe < 15: score += 10
        elif pe > 50: score -= 10

    # Clamp final score to 0-100 range
    return max(0, min(100, score))


def get_rank_folder(score):
    """Returns the rank folder name based on the score."""
    for folder_name, info in RANK_FOLDERS.items():
        if info["min"] <= score <= info["max"]:
            return folder_name
    return "Risky"


#---- Comparison Function ----

def compare_companies(data_list):
    
    #Compares financial data between multiple companies.

   # Takes a list of company data dictionaries and produces a comparison
    #showing each company's metrics side by side, along with averages
    #and which company ranks best/worst in each metric.
    
    if not data_list or len(data_list) < 2:
        return None

    comparison_metrics = [
        "Current Price", "Market Cap", "Total Revenue", "Gross Profits",
        "Net Income", "Total Cash", "Total Debt", "Book Value",
    ]
    ratio_metrics = [
        "Net Profit Margin", "Cash to Debt Ratio", "Market Cap to Revenue",
        "Debt to Equity", "Trailing PE", "Return on Equity",
    ]

    # Build comparison table
    #Had to use A.I in order to build the comparison table and make sure it calculates averages and handles missing values correctly.
    comparison = {}
    for metric in comparison_metrics:
        row = {}
        values = []
        for entry in data_list:
            val = safe_value(entry["data"].get(metric))
            ticker = entry["data"]["Ticker"]
            row[ticker] = val
            if val is not None:
                values.append(val)
        row["Average"] = sum(values) / len(values) if values else None
        comparison[metric] = row

    for metric in ratio_metrics:
        row = {}
        values = []
        for entry in data_list:
            val = safe_value(entry["ratios"].get(metric))
            ticker = entry["data"]["Ticker"]
            row[ticker] = val
            if val is not None:
                values.append(val)
        row["Average"] = sum(values) / len(values) if values else None
        comparison[metric] = row

    return comparison


#---- Chart Generation Functions ----
def make_chart(ticker_symbol, data, ratios):
    
    #Creates a simple bar chart of selected financial values.

    #The chart is saved into the static/charts folder so Flask can display it
    #on the results page.

#Important values to show on chart Revenue, Net Income, Cash, Debt. This is a simple way to visualize the company's financial health at a glance.
    chart_values = {
        "Revenue": safe_value(data.get("Total Revenue")),
        "Net Income": safe_value(data.get("Net Income")),
        "Cash": safe_value(data.get("Total Cash")),
        "Debt": safe_value(data.get("Total Debt")),
    }

    chart_values = {
        key: value for key, value in chart_values.items()
        if value is not None
    }

    if not chart_values:
        return None
#Color coding based on if number is negative=red or positive=green. Debt is always red since it's a liability.
    colors = []
    for key, value in chart_values.items():
        if key == "Debt":
            colors.append("#e74c3c")
        elif value >= 0:
            colors.append("#27ae60")
        else:
            colors.append("#e74c3c")

    plt.figure(figsize=(8, 5))
    plt.bar(chart_values.keys(), chart_values.values(), color=colors)
    plt.title(f"Selected Financial Data for {ticker_symbol}")

    ax = plt.gca()
    max_val = max(abs(v) for v in chart_values.values())
#Format y-axis labels based on the scale of the values (e.g., millions, billions) to keep the chart readable.
    if max_val >= 1_000_000_000:
        ax.yaxis.set_major_formatter(
            plt.FuncFormatter(lambda x, p: f"${x/1_000_000_000:.1f}B")
        )
        plt.ylabel("Amount (Billions USD)")
    elif max_val >= 1_000_000:
        ax.yaxis.set_major_formatter(
            plt.FuncFormatter(lambda x, p: f"${x/1_000_000:.1f}M")
        )
        plt.ylabel("Amount (Millions USD)")
    else:
        plt.ylabel("Amount (USD)")

    plt.xticks(rotation=20)
    plt.tight_layout()
#quick fix for when chart is not pathing correctly and that it doesn't show correctly

    chart_filename = f"{ticker_symbol}_financial_chart.png"
    chart_path = os.path.join(CHART_FOLDER, chart_filename)

    plt.savefig(chart_path)
    plt.close()

    return f"charts/{chart_filename}"


def make_comparison_chart(data_list):
    
    #Creates a comparison bar chart using company scores.
   # Saves the chart into static/charts so Flask can display it.
    
    if not data_list:
        return None

    tickers = []
    scores = []
#Compares the financial health scores of multiple companies and creates a bar chart to visualize the comparison.
    for entry in data_list:
        ticker = entry["data"].get("Ticker", "Unknown")

        ratios = entry.get("ratios", {})
        score = calculate_score(entry["data"], ratios)

        tickers.append(ticker)
        scores.append(score)

    plt.figure(figsize=(8, 5))
# Each bar is color-coded based on the score (green for high, yellow for medium, red for low) to make it easy to see which company is performing better at a glance.

    colors = []
    for score in scores:
        if score >= 75:
            colors.append("#27ae60")
        elif score >= 50:
            colors.append("#f1c40f")
        else:
            colors.append("#e74c3c")

#formating the chart with titles and labels to make it clear and easy to understand for the user.
    plt.bar(tickers, scores, color=colors)

    plt.title("Financial Health Score Comparison")
    plt.ylabel("Score")
    plt.ylim(0, 100)

    for i, score in enumerate(scores):
        plt.text(i, score + 2, str(score), ha="center")

    plt.tight_layout()
#pathing for chart png
    chart_filename = "comparison_chart.png"
    chart_path = os.path.join(CHART_FOLDER, chart_filename)

    plt.savefig(chart_path)
    plt.close()

    return f"charts/{chart_filename}"

#---- Output Saving Functions ----

def save_outputs(ticker_symbol, data, ratios, category, score, custom_filename=None):
    
    #Saves the results to both CSV and Excel.

   # This matches the proposal's export feature. The Excel file also includes basic
    #color formatting to make stronger and weaker values easier to notice.
    #Files are saved into rank-based subfolders.

#This puts it in the ranking folder based on the score and uses a a date format of year month day, then time
    rank_folder = get_rank_folder(score)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

#If User opts to name file, it will use that name instead of the default naming convention. It also replaces spaces with underscores to ensure the filename is valid and does not cause issues when saving or downloading.
    if custom_filename:
        base_name = custom_filename.replace(" ", "_")
    else:
        base_name = f"{ticker_symbol}_financial_analysis_{timestamp}"

    csv_filename = f"{base_name}.csv"
    excel_filename = f"{base_name}.xlsx"

    csv_path = os.path.join(OUTPUT_FOLDER, rank_folder, csv_filename)
    excel_path = os.path.join(OUTPUT_FOLDER, rank_folder, excel_filename)

    rows = []
    for key, value in data.items():
        rows.append({"Section": "Company Data", "Metric": key, "Value": value})
    for key, value in ratios.items():
        rows.append({"Section": "Calculated Ratios", "Metric": key, "Value": value})
    rows.append({"Section": "Category", "Metric": "Company Category", "Value": category})
    rows.append({"Section": "Score", "Metric": "Financial Health Score", "Value": score})

    df = pd.DataFrame(rows)
    df.to_csv(csv_path, index=False)

    # Create a styled Excel file with simple highlighting. Formatting for easier identification
    #Used A.I in order to find out how to manipulate/Add transfer data to excel file 
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Financial Analysis")
        worksheet = writer.sheets["Financial Analysis"]

        from openpyxl.styles import PatternFill, Font
#Adding color
#Used A.I in order to find patterns formatiting and how to add colors to excel file
#Also used A.I in rder to
        header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
        positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)

        for row in range(2, worksheet.max_row + 1):
            value_cell = worksheet.cell(row=row, column=3)
            try:
                value = float(value_cell.value)
                if value > 0:
                    value_cell.fill = positive_fill
                elif value < 0:
                    value_cell.fill = negative_fill
            except (TypeError, ValueError):
                pass

        worksheet.column_dimensions["A"].width = 22
        worksheet.column_dimensions["B"].width = 28
        worksheet.column_dimensions["C"].width = 24

# ---- Flask Routes ----
# These routes control the main pages and actions of the web application.
# Each route connects user actions from the website to the backend functions above.

    # Record to history of session with company name, ticker, category, score, and file paths. This allows the dashboard to show recent analyses and summary counts by rank.
    history = load_history()
    history.append({
       "username": session.get("username"),
        "ticker": ticker_symbol,
        "company_name": data.get("Company Name", ticker_symbol),
        "category": category,
        "score": score,
        "rank_folder": rank_folder,
        "csv_file": f"{rank_folder}/{csv_filename}",
        "excel_file": f"{rank_folder}/{excel_filename}",
        "timestamp": timestamp,
    })
    save_history(history)

    return f"{rank_folder}/{csv_filename}", f"{rank_folder}/{excel_filename}"


def get_value_class(key, value):
    
   #Returns a CSS class name for color highlighting table cells.

    #Green for positive/good values, red for negative/bad values.
   
    if value == "Not available":
        return "neutral"

    try:
        num = float(value.replace("$", "").replace(",", "").replace("%", ""))
    except (ValueError, TypeError, AttributeError):
        return "neutral"

    # Metrics where higher is better
    positive_metrics = [
        "Total Revenue", "Gross Profits", "Net Income", "Total Cash",
        "Book Value", "Current Price", "Market Cap",
        "Net Profit Margin", "Cash to Debt Ratio", "Return on Equity",
    ]
    # Metrics where lower is better
    negative_metrics = ["Total Debt", "Debt to Equity"]

    if key in positive_metrics:
        return "positive" if num > 0 else "negative"
    elif key in negative_metrics:
        return "negative" if num > 100 else "positive" if num >= 0 else "negative"
    elif key == "Trailing PE" or key == "Forward PE":
        if 0 < num < 25:
            return "positive"
        elif num > 40:
            return "negative"
        return "neutral"

    return "neutral"




# Dashboard route:
# Shows the logged-in user's recent analyses and summary counts and adds in the dashboard for watchlist
@app.route("/", methods=["GET"])
@login_required
def home():
    #Quick fox to rank_counts not being defined in local scope of dashboard route. Used A.I to hlep with this issue and how it can be defined
    global rank_counts
    #This is the main dashboard page that users see when they log in. It shows their recent analyses and summary counts of how many companies they've analyzed in each rank category. It also includes a watchlist widget that shows the tickers in their watchlist along with current price and score for each.
    username = session.get("username")
    all_history = load_history()
    history = [e for e in all_history if e.get("username") == username]
    recent = history[-5:][::-1]
    #Added in Watchlist to dashboard and each tied to personal users account 
    #Added company to watchlist widget that shows the current price and score for each ticker in the user's watchlist. This allows users to quickly see how their watchlist companies are doing without having to run a full analysis on each one.
    watchlist_widget = []
    users = load_users()

    user_data = users.get(username, {})

    # Convert old user format to new format if needed
    if isinstance(user_data, str):
        users[username] = {
            "password": user_data,
            "watchlist": []
        }
        save_users(users)
        user_data = users[username]

    watchlist_tickers = user_data.get("watchlist", [])
    for ticker in watchlist_tickers[:5]:
        try:
            data = get_financial_data(ticker)
            ratios = calculate_ratios(data)
            score = calculate_score(data, ratios)

            watchlist_widget.append({
                "ticker": ticker,
                "price": format_currency(data.get("Current Price")),
                "score": score,
                "rank": get_rank_folder(score),
            })

        except Exception:
            continue    
    
    #Count saved analyses by rank.
    rank_counts = {"High Rank": 0, "Stable": 0, "Risky": 0}
    for e in history:
        f = e.get("rank_folder")
        if f in rank_counts: rank_counts[f] += 1
    return render_template(
    "dashboard.html",
    recent=recent,
    rank_counts=rank_counts,
    total=len(history),
    watchlist_widget=watchlist_widget
)


# Watchlist route:
# Displays the logged-in user's saved watchlist and filtering options.
#This allows users to keep track of companies they are interested in without having to run a full analysis on each one. 
# They can see the current price and score for each ticker in their watchlist at a glance.
@app.route("/watchlist")
@login_required
def watchlist():
    username = session.get("username")
    users = load_users()

    # Handle older user format
    if username in users and isinstance(users[username], str):
        users[username] = {
            "password": users[username],
            "watchlist": []
        }
        save_users(users)

    watchlist_tickers = users.get(username, {}).get("watchlist", [])

    watchlist_data = []
#This display data for each ticker in user's watchlist
    for ticker in watchlist_tickers:
        try:
            data = get_financial_data(ticker)
            ratios = calculate_ratios(data)
            score = calculate_score(data, ratios)
#Appends title to each data collected
            watchlist_data.append({
                "ticker": ticker,
                "score": score,
                "profit": format_percent(ratios.get("Net Profit Margin")),
                "pe": format_number(ratios.get("Trailing PE")),
            })
#If no data found or entered ticker is invalid, it will skip that ticker and continue to the next one without crashing the page. This way users can still see valid tickers in their watchlist even if some are invalid or have missing data.
        except Exception:
            continue

    return render_template(
        "watchlist.html",
        watchlist=watchlist_data
    )


# Analyze route:
# Lets the user enter one ticker, runs the full analysis, saves outputs,
# and displays the results page.
@app.route("/analyze", methods=["GET", "POST"])
@login_required
def analyze():
    #Page where users enter a company ticker symbol for analysis.
    if request.method == "POST":
        ticker_symbol = request.form.get("ticker", "").upper().strip()
#error handling for empty or invalid ticker symbols. 
        if not ticker_symbol:
            return render_template("analyze.html", error="Please enter a ticker symbol.")

        username = session.get("username")
        add_to_watchlist(username, ticker_symbol)

        custom_filename = request.form.get("custom_filename", "").strip()
#This gets the finacial data and analysis that we did  such as ratios, category, sentimental analysis, and then saves the outputs
        try:
            data = get_financial_data(ticker_symbol)
            ratios = calculate_ratios(data)
            category = categorize_company(data, ratios)
            score = calculate_score(data, ratios)
            chart_file = make_chart(ticker_symbol, data, ratios)
            sentiment_label, sentiment_score, good, bad = get_sentiment_analysis(ticker_symbol)
            csv_file, excel_file = save_outputs(
                ticker_symbol, data, ratios, category, score,
                custom_filename=custom_filename if custom_filename else None
            )

            # Build display-ready data with CSS classes for color highlighting
            display_data = {}
            for key, value in data.items():
                if key in ["Current Price", "Market Cap", "Total Revenue", "Gross Profits",
                           "Net Income", "Total Cash", "Total Debt", "Book Value"]:
                    formatted = format_currency(value)
                else:
                    formatted = value if value is not None else "Not available"
                display_data[key] = {
                    "value": formatted,
                    "class": get_value_class(key, formatted),
                }
#Formating ratios 
            display_ratios = {}
            for key, value in ratios.items():
                formatted = format_number(value)
                display_ratios[key] = {
                    "value": formatted,
                    "class": get_value_class(key, formatted),
                }
#Retuns all the data to be displayed on the results page, including the company data, calculated ratios, category, score, chart file path, and sentiment analysis results. The template will use this data to show a comprehensive analysis of the company based on the entered ticker symbol.
            return render_template(
                "results.html",
                ticker=ticker_symbol,
                data=display_data,
                ratios=display_ratios,
                category=category,
                score=score,
                rank_folder=get_rank_folder(score),
                chart_file=chart_file,
                csv_file=csv_file,
                excel_file=excel_file,
                sentiment_label=sentiment_label,
                sentiment_score=sentiment_score,
                good_news=good,
                bad_news=bad,

            )
        except Exception as error:
            return render_template("analyze.html", error=str(error))

    return render_template("analyze.html")


# Compare route:
# Lets the user enter multiple tickers and compares valid companies side by side.
# Invalid tickers are skipped so the rest of the comparison can still run.
@app.route("/compare", methods=["GET", "POST"])
@login_required
def compare():
    #Industry Compare page for comparing multiple companies side by side.
    if request.method == "POST":
        tickers_raw = request.form.get("tickers", "")
        ticker_list = [t.strip().upper() for t in tickers_raw.replace(";", ",").split(",") if t.strip()]
#error handling for too few or too many ticker symbols entered. We require at least 2 for a comparison but limit to 5 to keep the results manageable and readable on the page.
        if len(ticker_list) < 2:
            return render_template("compare.html", error="Please enter at least 2 ticker symbols separated by commas.")

        if len(ticker_list) > 5:
            return render_template("compare.html", error="Please enter no more than 5 ticker symbols.")

        data_list = []
        invalid_tickers = []
#This loops through the entered ticker symbols, tries to get their financial data and calculate ratios, category, and score. I
# Invalid tickers are skipped so the rest of the comparison can still run.
        for ticker in ticker_list:
            try:
                data = get_financial_data(ticker)
                ratios = calculate_ratios(data)
                category = categorize_company(data, ratios)
                score = calculate_score(data, ratios)

                data_list.append({
                    "data": data,
                    "ratios": ratios,
                    "category": category,
                    "score": score,
                })

            except Exception:
                invalid_tickers.append(ticker)

        if len(data_list) < 2:
            return render_template(
                "compare.html",
                error="Please enter at least 2 valid ticker symbols."
            )

        comparison = compare_companies(data_list)
        chart_file = make_comparison_chart(data_list)

        currency_metrics = [
            "Current Price", "Market Cap", "Total Revenue", "Gross Profits",
            "Net Income", "Total Cash", "Total Debt", "Book Value",
        ]

        display_comparison = {}
        for metric, row in comparison.items():
            display_row = {}
            for col, val in row.items():
                if metric in currency_metrics:
                    formatted = format_currency(val)
                else:
                    formatted = format_number(val)

                display_row[col] = {
                    "value": formatted,
                    "class": get_value_class(metric, formatted),
                }

#This formats the comparison data for display in the template, including formatting numbers as currency where appropriate and adding CSS classes for color highlighting 
            display_comparison[metric] = display_row
        tickers = [entry["data"]["Ticker"] for entry in data_list]
        categories = {entry["data"]["Ticker"]: entry["category"] for entry in data_list}
        scores = {entry["data"]["Ticker"]: entry["score"] for entry in data_list}

        warning = None
        if len(invalid_tickers) > 0:
            warning = "Skipped invalid ticker(s): " + ", ".join(invalid_tickers)

#Results that are given after doing analysis for company comparison. 
        return render_template(
            "compare_results.html",
            tickers=tickers,
            comparison=display_comparison,
            categories=categories,
            scores=scores,
            chart_file=chart_file,
            warning=warning,
        )

    return render_template("compare.html")


# Downloads route:
# Shows saved CSV and Excel files for the logged-in user only.
@app.route("/downloads")
@login_required
def downloads():
    #Downloads Center page showing saved analyses for the logged-in user only.
#download files are attach to user and sessions
    username = session.get("username")
    history = [
        entry for entry in load_history()
        if entry.get("username") == username
    ]
#Basiclaly a reference to the rank folders and calls on the history of the users files
    files_by_rank = {}

    for folder_name, info in RANK_FOLDERS.items():
        files = []
#adds the files to download page for both csv and excel files 
        for entry in history:
            if entry.get("rank_folder") == folder_name:
                files.append({
                    "name": os.path.basename(entry.get("csv_file", "")),
                    "path": entry.get("csv_file", ""),
                    "size": "",
                    "ext": ".csv",
                })

                files.append({
                    "name": os.path.basename(entry.get("excel_file", "")),
                    "path": entry.get("excel_file", ""),
                    "size": "",
                    "ext": ".xlsx",
                })

        files_by_rank[folder_name] = {
            "label": info["label"],
            "files": files,
        }

    return render_template("downloads.html", files_by_rank=files_by_rank)


# File download route:
# Sends a selected CSV or Excel file to the user's browser.
@app.route("/download/<path:filename>")
@login_required
def download_file(filename):
    # Build full absolute path
    file_path = os.path.join(os.getcwd(), OUTPUT_FOLDER, filename)

    # Check if file exists
    if not os.path.exists(file_path):
        return f"Error: File not found on server: {file_path}", 404

    # Force browser download
    return send_file(
        file_path,
        as_attachment=True,
        download_name=os.path.basename(file_path)
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5001, debug=True)