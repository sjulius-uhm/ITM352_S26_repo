"""
Company Financial Data Web Scraper and Analysis Tool

AI use note:
This starter application was generated with AI support based on the project proposal.
The code was written to stay readable and class-friendly. Each major function has comments
explaining what the section does for the overall application.

Before submitting, review the code, test it, and make edits so you understand it fully.
"""

from flask import Flask, render_template, request, send_file, redirect, url_for, jsonify, session
from functools import wraps
import os
import json
from datetime import datetime

import pandas as pd
import matplotlib
matplotlib.use("Agg")  # Lets matplotlib save charts without opening a window
import matplotlib.pyplot as plt
import numpy as np

try:
    import yfinance as yf
except ImportError:
    yf = None


app = Flask(__name__)
app.secret_key = "fin-analytix-secret-key-2026"

USERS_FILE = "users.json"
OUTPUT_FOLDER = "outputs"
CHART_FOLDER = os.path.join("static", "charts")
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
os.makedirs(CHART_FOLDER, exist_ok=True)

WATCHLISTS_FILE = "watchlists.json"
MAX_WATCHLIST_SIZE = 20

# Rank folders for organizing saved analyses
RANK_FOLDERS = {
    "HIGH_RANK": {"label": "HIGH_RANK (Score 80-100)", "min": 80, "max": 100},
    "STABLE": {"label": "STABLE (Score 50-79)", "min": 50, "max": 79},
    "WATCHLIST": {"label": "WATCHLIST (Score 0-49)", "min": 0, "max": 49},
}

# Create rank subfolders inside outputs
for folder_name in RANK_FOLDERS:
    os.makedirs(os.path.join(OUTPUT_FOLDER, folder_name), exist_ok=True)

# File to persist analysis history
HISTORY_FILE = os.path.join(OUTPUT_FOLDER, "analysis_history.json")


def load_history():
    """Load analysis history from disk."""
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE, "r") as f:
            return json.load(f)
    return []


def save_history(history):
    """Save analysis history to disk."""
    with open(HISTORY_FILE, "w") as f:
        json.dump(history, f, indent=2)


def load_users():
    """Load user accounts from disk."""
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, "r") as f:
            return json.load(f)
    return {}


def save_users(users):
    """Save user accounts to disk."""
    with open(USERS_FILE, "w") as f:
        json.dump(users, f, indent=2)


def load_watchlists():
    if os.path.exists(WATCHLISTS_FILE):
        with open(WATCHLISTS_FILE, "r") as f:
            return json.load(f)
    return {}


def save_watchlists(watchlists):
    with open(WATCHLISTS_FILE, "w") as f:
        json.dump(watchlists, f, indent=2)


def get_user_watchlist(username):
    watchlists = load_watchlists()
    return watchlists.get(username, [])


def add_to_watchlist(username, ticker):
    watchlists = load_watchlists()
    user_list = watchlists.get(username, [])
    ticker = ticker.upper().strip()
    if ticker in user_list:
        return "already_exists"
    if len(user_list) >= MAX_WATCHLIST_SIZE:
        return "full"
    user_list.append(ticker)
    watchlists[username] = user_list
    save_watchlists(watchlists)
    return "added"


def remove_from_watchlist(username, ticker):
    watchlists = load_watchlists()
    user_list = watchlists.get(username, [])
    ticker = ticker.upper().strip()
    if ticker in user_list:
        user_list.remove(ticker)
        watchlists[username] = user_list
        save_watchlists(watchlists)


def get_watchlist_data(tickers):
    results = []
    for ticker in tickers:
        try:
            ticker = ticker.upper().strip()
            company = yf.Ticker(ticker)
            info = company.info
            if not info or not info.get("currentPrice"):
                continue

            current = info.get("currentPrice") or info.get("regularMarketPrice")
            previous = info.get("previousClose")
            if current and previous and previous != 0:
                pct_change = ((current - previous) / previous) * 100
            else:
                pct_change = None

            data = get_financial_data(ticker)
            ratios = calculate_ratios(data)
            category = categorize_company(data, ratios)
            score = calculate_score(data, ratios)

            results.append({
                "ticker": ticker,
                "company_name": info.get("longName", ticker),
                "current_price": current,
                "pct_change": pct_change,
                "profit_margin": info.get("profitMargins"),
                "debt_to_equity": info.get("debtToEquity"),
                "trailing_pe": info.get("trailingPE"),
                "revenue": info.get("totalRevenue"),
                "score": score,
                "category": category,
            })
        except Exception:
            continue
    return results


def login_required(f):
    """Decorator that redirects to login page if user is not signed in."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if "username" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


@app.route("/login", methods=["GET", "POST"])
def login():
    """Login page."""
    if "username" in session:
        return redirect(url_for("home"))

    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        if not username or not password:
            error = "Please fill in both fields."
        else:
            users = load_users()
            if username in users and users[username] == password:
                session["username"] = username
                return redirect(url_for("home"))
            else:
                error = "Invalid username or password."

    return render_template("login.html", error=error)


@app.route("/register", methods=["GET", "POST"])
def register():
    """Registration page."""
    if "username" in session:
        return redirect(url_for("home"))

    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        confirm = request.form.get("confirm", "").strip()

        if not username or not password:
            error = "Please fill in all fields."
        elif len(password) < 4:
            error = "Password must be at least 4 characters."
        elif password != confirm:
            error = "Passwords do not match."
        else:
            users = load_users()
            if username in users:
                error = "That username is already taken."
            else:
                users[username] = password
                save_users(users)
                session["username"] = username
                return redirect(url_for("home"))

    return render_template("register.html", error=error)


@app.route("/logout")
def logout():
    """Log the user out."""
    session.pop("username", None)
    return redirect(url_for("login"))


def safe_value(value):
    """Converts missing or invalid values into None so the app does not crash."""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except TypeError:
        pass
    return value


def format_currency(value):
    """Formats large financial numbers for display on the webpage."""
    value = safe_value(value)
    if value is None:
        return "Not available"
    try:
        return "${:,.0f}".format(float(value))
    except (ValueError, TypeError):
        return "Not available"


def format_number(value):
    """Formats ratios and percentages for display."""
    value = safe_value(value)
    if value is None:
        return "Not available"
    try:
        return "{:,.2f}".format(float(value))
    except (ValueError, TypeError):
        return "Not available"


def format_percent(value):
    """Formats a decimal ratio as a percentage string."""
    value = safe_value(value)
    if value is None:
        return "Not available"
    try:
        return "{:.2f}%".format(float(value) * 100)
    except (ValueError, TypeError):
        return "Not available"


def clean_data(data):
    """
    Processes raw data by handling missing values and ensuring consistent
    formatting for analysis.

    This function normalizes the raw dictionary returned by get_financial_data()
    so downstream functions receive clean, predictable values.
    """
    cleaned = {}
    for key, value in data.items():
        value = safe_value(value)

        # Ensure numeric fields are proper floats
        numeric_fields = [
            "Current Price", "Market Cap", "Total Revenue", "Gross Profits",
            "Net Income", "Total Cash", "Total Debt", "Book Value",
            "Profit Margin", "Return on Equity", "Debt to Equity",
            "Trailing PE", "Forward PE",
        ]
        if key in numeric_fields and value is not None:
            try:
                value = float(value)
            except (ValueError, TypeError):
                value = None

        # Ensure string fields are stripped strings
        string_fields = ["Ticker", "Company Name", "Sector", "Industry", "Recommendation"]
        if key in string_fields and value is not None:
            value = str(value).strip()

        cleaned[key] = value

    return cleaned


def get_financial_data(ticker_symbol):
    """
    Retrieves company financial data using yfinance.

    The original proposal mentions Yahoo Finance, SEC, requests, and BeautifulSoup.
    This version uses yfinance because it is more stable for a student project than scraping
    website HTML directly. The rest of the app still follows the proposal: collect data,
    organize it, analyze it, export it, and display it through Flask.
    """
    if yf is None:
        raise RuntimeError("yfinance is not installed. Run: pip install yfinance")

    ticker_symbol = ticker_symbol.upper().strip()
    company = yf.Ticker(ticker_symbol)
    info = company.info

    if not info or info.get("regularMarketPrice") is None and info.get("currentPrice") is None:
        raise ValueError("No company data was found. Check the ticker symbol and try again.")

    # Pull important financial values from yfinance's company information dictionary.
    data = {
        "Ticker": ticker_symbol,
        "Company Name": info.get("longName"),
        "Sector": info.get("sector"),
        "Industry": info.get("industry"),
        "Current Price": info.get("currentPrice") or info.get("regularMarketPrice"),
        "Market Cap": info.get("marketCap"),
        "Total Revenue": info.get("totalRevenue"),
        "Gross Profits": info.get("grossProfits"),
        "Net Income": info.get("netIncomeToCommon"),
        "Total Cash": info.get("totalCash"),
        "Total Debt": info.get("totalDebt"),
        "Book Value": info.get("bookValue"),
        "Profit Margin": info.get("profitMargins"),
        "Return on Equity": info.get("returnOnEquity"),
        "Debt to Equity": info.get("debtToEquity"),
        "Trailing PE": info.get("trailingPE"),
        "Forward PE": info.get("forwardPE"),
        "Recommendation": info.get("recommendationKey"),
    }

    return clean_data(data)


def calculate_ratios(data):
    """
    Performs simple financial calculations.

    These calculations support the proposal's goal of turning raw financial data
    into useful decision-making information.
    """
    total_revenue = safe_value(data.get("Total Revenue"))
    net_income = safe_value(data.get("Net Income"))
    total_cash = safe_value(data.get("Total Cash"))
    total_debt = safe_value(data.get("Total Debt"))
    market_cap = safe_value(data.get("Market Cap"))

    ratios = {}

    if total_revenue and net_income is not None and total_revenue != 0:
        ratios["Net Profit Margin"] = net_income / total_revenue
    else:
        ratios["Net Profit Margin"] = None

    if total_debt and total_debt != 0 and total_cash is not None:
        ratios["Cash to Debt Ratio"] = total_cash / total_debt
    else:
        ratios["Cash to Debt Ratio"] = None

    if total_revenue and total_revenue != 0 and market_cap is not None:
        ratios["Market Cap to Revenue"] = market_cap / total_revenue
    else:
        ratios["Market Cap to Revenue"] = None

    ratios["Debt to Equity"] = safe_value(data.get("Debt to Equity"))
    ratios["Trailing PE"] = safe_value(data.get("Trailing PE"))
    ratios["Return on Equity"] = safe_value(data.get("Return on Equity"))

    return ratios


def categorize_company(data, ratios):
    """
    Places the company into a simple category based on defined criteria.

    This matches the proposal feature of sorting companies into categories such as
    growth, value, or losing companies. These rules are simple and can be adjusted.
    """
    net_income = safe_value(data.get("Net Income"))
    revenue = safe_value(data.get("Total Revenue"))
    pe_ratio = safe_value(ratios.get("Trailing PE"))
    profit_margin = safe_value(ratios.get("Net Profit Margin"))

    if net_income is not None and net_income < 0:
        return "Losing Company"

    if revenue and profit_margin is not None and profit_margin > 0.15:
        return "Growth / Strong Profit Company"

    if pe_ratio is not None and 0 < pe_ratio < 20:
        return "Possible Value Company"

    return "Neutral / Needs More Review"


def calculate_score(data, ratios):
    """
    Calculates a simple financial health score from 0-100.

    This score is used to sort companies into rank folders
    (HIGH_RANK 80-100, STABLE 50-79, WATCHLIST 0-49).
    """
    score = 50  # Start at neutral

    net_income = safe_value(data.get("Net Income"))
    profit_margin = safe_value(ratios.get("Net Profit Margin"))
    roe = safe_value(ratios.get("Return on Equity"))
    de_ratio = safe_value(ratios.get("Debt to Equity"))
    pe_ratio = safe_value(ratios.get("Trailing PE"))
    cash_to_debt = safe_value(ratios.get("Cash to Debt Ratio"))

    # Profitability
    if net_income is not None:
        if net_income > 0:
            score += 10
        else:
            score -= 20

    if profit_margin is not None:
        if profit_margin > 0.20:
            score += 15
        elif profit_margin > 0.10:
            score += 10
        elif profit_margin > 0:
            score += 5
        else:
            score -= 10

    # Return on equity
    if roe is not None:
        if roe > 0.20:
            score += 10
        elif roe > 0.10:
            score += 5
        elif roe < 0:
            score -= 10

    # Debt management
    if de_ratio is not None:
        if de_ratio < 50:
            score += 10
        elif de_ratio < 100:
            score += 5
        elif de_ratio > 200:
            score -= 10

    # Cash position
    if cash_to_debt is not None:
        if cash_to_debt > 1:
            score += 5
        elif cash_to_debt < 0.2:
            score -= 5

    # Valuation
    if pe_ratio is not None:
        if 0 < pe_ratio < 15:
            score += 5
        elif 15 <= pe_ratio < 25:
            score += 2
        elif pe_ratio > 50:
            score -= 5

    # Clamp to 0-100
    return max(0, min(100, score))


def get_rank_folder(score):
    """Returns the rank folder name based on the score."""
    for folder_name, info in RANK_FOLDERS.items():
        if info["min"] <= score <= info["max"]:
            return folder_name
    return "WATCHLIST"


def compare_companies(data_list):
    """
    Compares financial data between multiple companies.

    Takes a list of company data dictionaries and produces a comparison
    showing each company's metrics side by side, along with averages
    and which company ranks best/worst in each metric.
    """
    if not data_list or len(data_list) < 2:
        return None

    comparison_metrics = [
        "Current Price", "Market Cap", "Total Revenue", "Gross Profits",
        "Net Income", "Total Cash", "Total Debt", "Book Value",
    ]
    ratio_metrics = [
        "Net Profit Margin", "Cash to Debt Ratio", "Market Cap to Revenue",
        "Debt to Equity", "Trailing PE", "Return on Equity",
    ]

    # Build comparison table
    comparison = {}
    for metric in comparison_metrics:
        row = {}
        values = []
        for entry in data_list:
            val = safe_value(entry["data"].get(metric))
            ticker = entry["data"]["Ticker"]
            row[ticker] = val
            if val is not None:
                values.append(val)
        row["Average"] = sum(values) / len(values) if values else None
        comparison[metric] = row

    for metric in ratio_metrics:
        row = {}
        values = []
        for entry in data_list:
            val = safe_value(entry["ratios"].get(metric))
            ticker = entry["data"]["Ticker"]
            row[ticker] = val
            if val is not None:
                values.append(val)
        row["Average"] = sum(values) / len(values) if values else None
        comparison[metric] = row

    return comparison


def make_chart(ticker_symbol, data, ratios):
    """
    Creates a simple bar chart of selected financial values.

    The chart is saved into the static folder so Flask can display it on the results page.
    """
    chart_values = {
        "Revenue": safe_value(data.get("Total Revenue")),
        "Net Income": safe_value(data.get("Net Income")),
        "Cash": safe_value(data.get("Total Cash")),
        "Debt": safe_value(data.get("Total Debt")),
    }

    # Remove values that are missing so the chart only shows available data.
    chart_values = {key: value for key, value in chart_values.items() if value is not None}

    if not chart_values:
        return None

    colors = []
    for key, val in chart_values.items():
        if key == "Debt":
            colors.append("#e74c3c")
        elif val >= 0:
            colors.append("#27ae60")
        else:
            colors.append("#e74c3c")

    plt.figure(figsize=(8, 5))
    plt.bar(chart_values.keys(), chart_values.values(), color=colors)
    plt.title(f"Selected Financial Data for {ticker_symbol}")

    # Format y-axis to show billions/millions instead of raw numbers
    ax = plt.gca()
    max_val = max(abs(v) for v in chart_values.values())
    if max_val >= 1_000_000_000:
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f"${x/1_000_000_000:.1f}B"))
        plt.ylabel("Amount (Billions USD)")
    elif max_val >= 1_000_000:
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f"${x/1_000_000:.1f}M"))
        plt.ylabel("Amount (Millions USD)")
    else:
        plt.ylabel("Amount (USD)")

    plt.xticks(rotation=20)
    plt.tight_layout()

    chart_filename = f"{ticker_symbol}_financial_chart.png"
    chart_path = os.path.join(CHART_FOLDER, chart_filename)
    plt.savefig(chart_path)
    plt.close()

    return f"charts/{chart_filename}"


def make_comparison_chart(data_list):
    """
    Creates a grouped bar chart comparing multiple companies.
    """
    tickers = [entry["data"]["Ticker"] for entry in data_list]
    metrics = ["Total Revenue", "Net Income", "Total Cash", "Total Debt"]

    values_by_metric = {}
    for metric in metrics:
        vals = []
        for entry in data_list:
            v = safe_value(entry["data"].get(metric))
            vals.append(v if v is not None else 0)
        values_by_metric[metric] = vals

    x = np.arange(len(tickers))
    width = 0.2
    fig, ax = plt.subplots(figsize=(10, 6))

    for i, (metric, vals) in enumerate(values_by_metric.items()):
        ax.bar(x + i * width, vals, width, label=metric)

    ax.set_xlabel("Company")
    ax.set_title("Company Comparison")
    ax.set_xticks(x + width * 1.5)
    ax.set_xticklabels(tickers)
    ax.legend()

    # Format y-axis to show billions/millions
    all_vals = [v for vals in values_by_metric.values() for v in vals if v != 0]
    if all_vals:
        max_val = max(abs(v) for v in all_vals)
        if max_val >= 1_000_000_000:
            ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f"${x/1_000_000_000:.1f}B"))
            ax.set_ylabel("Amount (Billions USD)")
        elif max_val >= 1_000_000:
            ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f"${x/1_000_000:.1f}M"))
            ax.set_ylabel("Amount (Millions USD)")
        else:
            ax.set_ylabel("Amount (USD)")
    else:
        ax.set_ylabel("Amount (USD)")

    plt.tight_layout()

    chart_filename = "comparison_chart.png"
    chart_path = os.path.join(CHART_FOLDER, chart_filename)
    plt.savefig(chart_path)
    plt.close()

    return f"charts/{chart_filename}"


def save_outputs(ticker_symbol, data, ratios, category, score, custom_filename=None):
    """
    Saves the results to both CSV and Excel.

    This matches the proposal's export feature. The Excel file also includes basic
    color formatting to make stronger and weaker values easier to notice.
    Files are saved into rank-based subfolders.
    """
    rank_folder = get_rank_folder(score)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if custom_filename:
        base_name = custom_filename.replace(" ", "_")
    else:
        base_name = f"{ticker_symbol}_financial_analysis_{timestamp}"

    csv_filename = f"{base_name}.csv"
    excel_filename = f"{base_name}.xlsx"

    csv_path = os.path.join(OUTPUT_FOLDER, rank_folder, csv_filename)
    excel_path = os.path.join(OUTPUT_FOLDER, rank_folder, excel_filename)

    rows = []
    for key, value in data.items():
        rows.append({"Section": "Company Data", "Metric": key, "Value": value})
    for key, value in ratios.items():
        rows.append({"Section": "Calculated Ratios", "Metric": key, "Value": value})
    rows.append({"Section": "Category", "Metric": "Company Category", "Value": category})
    rows.append({"Section": "Score", "Metric": "Financial Health Score", "Value": score})

    df = pd.DataFrame(rows)
    df.to_csv(csv_path, index=False)

    # Create a styled Excel file with simple highlighting.
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Financial Analysis")
        workbook = writer.book
        worksheet = writer.sheets["Financial Analysis"]

        from openpyxl.styles import PatternFill, Font

        header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
        positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)

        for row in range(2, worksheet.max_row + 1):
            value_cell = worksheet.cell(row=row, column=3)
            try:
                value = float(value_cell.value)
                if value > 0:
                    value_cell.fill = positive_fill
                elif value < 0:
                    value_cell.fill = negative_fill
            except (TypeError, ValueError):
                pass

        worksheet.column_dimensions["A"].width = 22
        worksheet.column_dimensions["B"].width = 28
        worksheet.column_dimensions["C"].width = 24

    # Record to history
    history = load_history()
    history.append({
        "ticker": ticker_symbol,
        "company_name": data.get("Company Name", ticker_symbol),
        "category": category,
        "score": score,
        "rank_folder": rank_folder,
        "csv_file": f"{rank_folder}/{csv_filename}",
        "excel_file": f"{rank_folder}/{excel_filename}",
        "timestamp": timestamp,
    })
    save_history(history)

    return f"{rank_folder}/{csv_filename}", f"{rank_folder}/{excel_filename}"


def get_value_class(key, value):
    """
    Returns a CSS class name for color highlighting table cells.

    Green for positive/good values, red for negative/bad values.
    This implements the color highlighting from Milestone 3.
    """
    if value == "Not available":
        return "neutral"

    try:
        num = float(value.replace("$", "").replace(",", "").replace("%", ""))
    except (ValueError, TypeError, AttributeError):
        return "neutral"

    # Metrics where higher is better
    positive_metrics = [
        "Total Revenue", "Gross Profits", "Net Income", "Total Cash",
        "Book Value", "Current Price", "Market Cap",
        "Net Profit Margin", "Cash to Debt Ratio", "Return on Equity",
    ]
    # Metrics where lower is better
    negative_metrics = ["Total Debt", "Debt to Equity"]

    if key in positive_metrics:
        return "positive" if num > 0 else "negative"
    elif key in negative_metrics:
        return "negative" if num > 100 else "positive" if num >= 0 else "negative"
    elif key == "Trailing PE" or key == "Forward PE":
        if 0 < num < 25:
            return "positive"
        elif num > 40:
            return "negative"
        return "neutral"

    return "neutral"


# ---- Flask Routes ----

@app.route("/", methods=["GET"])
@login_required
def home():
    """Dashboard / home page showing recent analyses and quick stats."""
    history = load_history()
    recent = history[-5:][::-1] if history else []

    # Count by rank
    rank_counts = {"HIGH_RANK": 0, "STABLE": 0, "WATCHLIST": 0}
    for entry in history:
        folder = entry.get("rank_folder", "WATCHLIST")
        if folder in rank_counts:
            rank_counts[folder] += 1

    # Fetch top movers from user's watchlist
    top_movers = []
    tickers = get_user_watchlist(session["username"])
    if tickers:
        all_data = get_watchlist_data(tickers)
        movers_with_change = [s for s in all_data if s["pct_change"] is not None]
        movers_with_change.sort(key=lambda s: abs(s["pct_change"]), reverse=True)
        top_movers = movers_with_change[:5]

    return render_template(
        "dashboard.html",
        recent=recent,
        rank_counts=rank_counts,
        total=len(history),
        top_movers=top_movers,
    )


@app.route("/analyze", methods=["GET", "POST"])
@login_required
def analyze():
    """Page where users enter a company ticker symbol for analysis."""
    if request.method == "POST":
        ticker_symbol = request.form.get("ticker", "").upper().strip()
        if not ticker_symbol:
            return render_template("analyze.html", error="Please enter a ticker symbol.")

        custom_filename = request.form.get("custom_filename", "").strip()

        try:
            data = get_financial_data(ticker_symbol)
            ratios = calculate_ratios(data)
            category = categorize_company(data, ratios)
            score = calculate_score(data, ratios)
            chart_file = make_chart(ticker_symbol, data, ratios)
            csv_file, excel_file = save_outputs(
                ticker_symbol, data, ratios, category, score,
                custom_filename=custom_filename if custom_filename else None
            )

            # Build display-ready data with CSS classes for color highlighting
            display_data = {}
            for key, value in data.items():
                if key in ["Current Price", "Market Cap", "Total Revenue", "Gross Profits",
                           "Net Income", "Total Cash", "Total Debt", "Book Value"]:
                    formatted = format_currency(value)
                else:
                    formatted = value if value is not None else "Not available"
                display_data[key] = {
                    "value": formatted,
                    "class": get_value_class(key, formatted),
                }

            display_ratios = {}
            for key, value in ratios.items():
                formatted = format_number(value)
                display_ratios[key] = {
                    "value": formatted,
                    "class": get_value_class(key, formatted),
                }

            in_watchlist = ticker_symbol in get_user_watchlist(session["username"])

            return render_template(
                "results.html",
                ticker=ticker_symbol,
                data=display_data,
                ratios=display_ratios,
                category=category,
                score=score,
                rank_folder=get_rank_folder(score),
                chart_file=chart_file,
                csv_file=csv_file,
                excel_file=excel_file,
                in_watchlist=in_watchlist,
            )
        except Exception as error:
            return render_template("analyze.html", error=str(error))

    return render_template("analyze.html")


@app.route("/compare", methods=["GET", "POST"])
@login_required
def compare():
    """Industry Compare page for comparing multiple companies side by side."""
    if request.method == "POST":
        tickers_raw = request.form.get("tickers", "")
        ticker_list = [t.strip().upper() for t in tickers_raw.replace(";", ",").split(",") if t.strip()]

        if len(ticker_list) < 2:
            return render_template("compare.html", error="Please enter at least 2 ticker symbols separated by commas.")

        if len(ticker_list) > 5:
            return render_template("compare.html", error="Please enter no more than 5 ticker symbols.")

        try:
            data_list = []
            for ticker in ticker_list:
                data = get_financial_data(ticker)
                ratios = calculate_ratios(data)
                category = categorize_company(data, ratios)
                score = calculate_score(data, ratios)
                data_list.append({
                    "data": data,
                    "ratios": ratios,
                    "category": category,
                    "score": score,
                })

            comparison = compare_companies(data_list)
            chart_file = make_comparison_chart(data_list)

            # Format comparison values for display
            currency_metrics = [
                "Current Price", "Market Cap", "Total Revenue", "Gross Profits",
                "Net Income", "Total Cash", "Total Debt", "Book Value",
            ]
            display_comparison = {}
            for metric, row in comparison.items():
                display_row = {}
                for col, val in row.items():
                    if metric in currency_metrics:
                        formatted = format_currency(val)
                    else:
                        formatted = format_number(val)
                    display_row[col] = {
                        "value": formatted,
                        "class": get_value_class(metric, formatted),
                    }
                display_comparison[metric] = display_row

            tickers = [entry["data"]["Ticker"] for entry in data_list]
            categories = {entry["data"]["Ticker"]: entry["category"] for entry in data_list}
            scores = {entry["data"]["Ticker"]: entry["score"] for entry in data_list}

            return render_template(
                "compare_results.html",
                tickers=tickers,
                comparison=display_comparison,
                categories=categories,
                scores=scores,
                chart_file=chart_file,
            )
        except Exception as error:
            return render_template("compare.html", error=str(error))

    return render_template("compare.html")


@app.route("/downloads")
@login_required
def downloads():
    """Downloads Center page showing saved analyses organized by rank folder."""
    files_by_rank = {}
    for folder_name, info in RANK_FOLDERS.items():
        folder_path = os.path.join(OUTPUT_FOLDER, folder_name)
        files = []
        if os.path.exists(folder_path):
            for fname in sorted(os.listdir(folder_path), reverse=True):
                fpath = os.path.join(folder_path, fname)
                if os.path.isfile(fpath):
                    files.append({
                        "name": fname,
                        "path": f"{folder_name}/{fname}",
                        "size": os.path.getsize(fpath),
                        "ext": os.path.splitext(fname)[1].lower(),
                    })
        files_by_rank[folder_name] = {
            "label": info["label"],
            "files": files,
        }

    return render_template("downloads.html", files_by_rank=files_by_rank)


@app.route("/watchlist/add", methods=["POST"])
@login_required
def watchlist_add():
    ticker = request.form.get("ticker", "").upper().strip()
    if not ticker:
        return redirect(request.referrer or url_for("home"))
    add_to_watchlist(session["username"], ticker)
    return redirect(request.referrer or url_for("home"))


@app.route("/watchlist/remove", methods=["POST"])
@login_required
def watchlist_remove():
    ticker = request.form.get("ticker", "").upper().strip()
    if ticker:
        remove_from_watchlist(session["username"], ticker)
    return redirect(url_for("watchlist"))


FILTER_PRESETS = {
    "overall": {"key": "score", "reverse": True, "label": "Overall"},
    "profitability": {"key": "profit_margin", "reverse": True, "label": "Profitability"},
    "low_debt": {"key": "debt_to_equity", "reverse": False, "label": "Low Debt"},
    "value": {"key": "trailing_pe", "reverse": False, "label": "Value"},
    "growth": {"key": "revenue", "reverse": True, "label": "Growth"},
}


@app.route("/watchlist")
@login_required
def watchlist():
    tickers = get_user_watchlist(session["username"])
    active_filter = request.args.get("filter", "overall")
    if active_filter not in FILTER_PRESETS:
        active_filter = "overall"

    if not tickers:
        return render_template("watchlist.html", stocks=[], filters=FILTER_PRESETS, active_filter=active_filter)

    stocks = get_watchlist_data(tickers)

    preset = FILTER_PRESETS[active_filter]
    sort_key = preset["key"]

    def sort_val(item):
        v = item.get(sort_key)
        if v is None:
            return float('inf') if not preset["reverse"] else float('-inf')
        return v

    stocks.sort(key=sort_val, reverse=preset["reverse"])

    return render_template("watchlist.html", stocks=stocks, filters=FILTER_PRESETS, active_filter=active_filter)


@app.route("/download/<path:filename>")
@login_required
def download_file(filename):
    """Lets users download the CSV or Excel file created by the app."""
    file_path = os.path.join(OUTPUT_FOLDER, filename)
    # Prevent path traversal attacks
    real_output = os.path.realpath(OUTPUT_FOLDER)
    real_file = os.path.realpath(file_path)
    if not real_file.startswith(real_output):
        return "Access denied", 403
    if not os.path.exists(file_path):
        return "File not found", 404
    return send_file(file_path, as_attachment=True)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
