"""
Company Financial Data Web Scraper and Analysis Tool

AI use note:
This starter application was generated with AI support based on the project proposal.
The code was written to stay readable and class-friendly. Each major function has comments
explaining what the section does for the overall application.

Before submitting, review the code, test it, and make edits so you understand it fully.
"""

from flask import Flask, render_template, request, send_file
import os
from datetime import datetime

import pandas as pd
import matplotlib
matplotlib.use("Agg")  # Lets matplotlib save charts without opening a window
import matplotlib.pyplot as plt

try:
    import yfinance as yf
except ImportError:
    yf = None


app = Flask(__name__)

OUTPUT_FOLDER = "outputs"
CHART_FOLDER = os.path.join("static", "charts")
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
os.makedirs(CHART_FOLDER, exist_ok=True)


def safe_value(value):
    """Converts missing or invalid values into None so the app does not crash."""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except TypeError:
        pass
    return value


def format_currency(value):
    """Formats large financial numbers for display on the webpage."""
    value = safe_value(value)
    if value is None:
        return "Not available"
    try:
        return "${:,.0f}".format(float(value))
    except (ValueError, TypeError):
        return "Not available"


def format_number(value):
    """Formats ratios and percentages for display."""
    value = safe_value(value)
    if value is None:
        return "Not available"
    try:
        return "{:,.2f}".format(float(value))
    except (ValueError, TypeError):
        return "Not available"


def get_financial_data(ticker_symbol):
    """
    Retrieves company financial data using yfinance.

    The original proposal mentions Yahoo Finance, SEC, requests, and BeautifulSoup.
    This version uses yfinance because it is more stable for a student project than scraping
    website HTML directly. The rest of the app still follows the proposal: collect data,
    organize it, analyze it, export it, and display it through Flask.
    """
    if yf is None:
        raise RuntimeError("yfinance is not installed. Run: pip install yfinance")

    ticker_symbol = ticker_symbol.upper().strip()
    company = yf.Ticker(ticker_symbol)
    info = company.info

    if not info or info.get("regularMarketPrice") is None and info.get("currentPrice") is None:
        raise ValueError("No company data was found. Check the ticker symbol and try again.")

    # Pull important financial values from yfinance's company information dictionary.
    data = {
        "Ticker": ticker_symbol,
        "Company Name": info.get("longName"),
        "Sector": info.get("sector"),
        "Industry": info.get("industry"),
        "Current Price": info.get("currentPrice") or info.get("regularMarketPrice"),
        "Market Cap": info.get("marketCap"),
        "Total Revenue": info.get("totalRevenue"),
        "Gross Profits": info.get("grossProfits"),
        "Net Income": info.get("netIncomeToCommon"),
        "Total Cash": info.get("totalCash"),
        "Total Debt": info.get("totalDebt"),
        "Book Value": info.get("bookValue"),
        "Profit Margin": info.get("profitMargins"),
        "Return on Equity": info.get("returnOnEquity"),
        "Debt to Equity": info.get("debtToEquity"),
        "Trailing PE": info.get("trailingPE"),
        "Forward PE": info.get("forwardPE"),
        "Recommendation": info.get("recommendationKey"),
    }

    return data


def calculate_ratios(data):
    """
    Performs simple financial calculations.

    These calculations support the proposal's goal of turning raw financial data
    into useful decision-making information.
    """
    total_revenue = safe_value(data.get("Total Revenue"))
    net_income = safe_value(data.get("Net Income"))
    total_cash = safe_value(data.get("Total Cash"))
    total_debt = safe_value(data.get("Total Debt"))
    market_cap = safe_value(data.get("Market Cap"))

    ratios = {}

    if total_revenue and net_income is not None and total_revenue != 0:
        ratios["Net Profit Margin"] = net_income / total_revenue
    else:
        ratios["Net Profit Margin"] = None

    if total_debt and total_debt != 0 and total_cash is not None:
        ratios["Cash to Debt Ratio"] = total_cash / total_debt
    else:
        ratios["Cash to Debt Ratio"] = None

    if total_revenue and total_revenue != 0 and market_cap is not None:
        ratios["Market Cap to Revenue"] = market_cap / total_revenue
    else:
        ratios["Market Cap to Revenue"] = None

    ratios["Debt to Equity"] = safe_value(data.get("Debt to Equity"))
    ratios["Trailing PE"] = safe_value(data.get("Trailing PE"))
    ratios["Return on Equity"] = safe_value(data.get("Return on Equity"))

    return ratios


def categorize_company(data, ratios):
    """
    Places the company into a simple category based on defined criteria.

    This matches the proposal feature of sorting companies into categories such as
    growth, value, or losing companies. These rules are simple and can be adjusted.
    """
    net_income = safe_value(data.get("Net Income"))
    revenue = safe_value(data.get("Total Revenue"))
    pe_ratio = safe_value(ratios.get("Trailing PE"))
    profit_margin = safe_value(ratios.get("Net Profit Margin"))

    if net_income is not None and net_income < 0:
        return "Losing Company"

    if revenue and profit_margin is not None and profit_margin > 0.15:
        return "Growth / Strong Profit Company"

    if pe_ratio is not None and 0 < pe_ratio < 20:
        return "Possible Value Company"

    return "Neutral / Needs More Review"


def make_chart(ticker_symbol, data, ratios):
    """
    Creates a simple bar chart of selected financial values.

    The chart is saved into the static folder so Flask can display it on the results page.
    """
    chart_values = {
        "Revenue": safe_value(data.get("Total Revenue")),
        "Net Income": safe_value(data.get("Net Income")),
        "Cash": safe_value(data.get("Total Cash")),
        "Debt": safe_value(data.get("Total Debt")),
    }

    # Remove values that are missing so the chart only shows available data.
    chart_values = {key: value for key, value in chart_values.items() if value is not None}

    if not chart_values:
        return None

    plt.figure(figsize=(8, 5))
    plt.bar(chart_values.keys(), chart_values.values())
    plt.title(f"Selected Financial Data for {ticker_symbol}")
    plt.ylabel("Amount in Dollars")
    plt.xticks(rotation=20)
    plt.tight_layout()

    chart_filename = f"{ticker_symbol}_financial_chart.png"
    chart_path = os.path.join(CHART_FOLDER, chart_filename)
    plt.savefig(chart_path)
    plt.close()

    return f"charts/{chart_filename}"


def save_outputs(ticker_symbol, data, ratios, category):
    """
    Saves the results to both CSV and Excel.

    This matches the proposal's export feature. The Excel file also includes basic
    color formatting to make stronger and weaker values easier to notice.
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_filename = f"{ticker_symbol}_financial_analysis_{timestamp}.csv"
    excel_filename = f"{ticker_symbol}_financial_analysis_{timestamp}.xlsx"

    csv_path = os.path.join(OUTPUT_FOLDER, csv_filename)
    excel_path = os.path.join(OUTPUT_FOLDER, excel_filename)

    rows = []
    for key, value in data.items():
        rows.append({"Section": "Company Data", "Metric": key, "Value": value})
    for key, value in ratios.items():
        rows.append({"Section": "Calculated Ratios", "Metric": key, "Value": value})
    rows.append({"Section": "Category", "Metric": "Company Category", "Value": category})

    df = pd.DataFrame(rows)
    df.to_csv(csv_path, index=False)

    # Create a styled Excel file with simple highlighting.
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Financial Analysis")
        workbook = writer.book
        worksheet = writer.sheets["Financial Analysis"]

        from openpyxl.styles import PatternFill, Font

        header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
        positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)

        for row in range(2, worksheet.max_row + 1):
            value_cell = worksheet.cell(row=row, column=3)
            try:
                value = float(value_cell.value)
                if value > 0:
                    value_cell.fill = positive_fill
                elif value < 0:
                    value_cell.fill = negative_fill
            except (TypeError, ValueError):
                pass

        worksheet.column_dimensions["A"].width = 22
        worksheet.column_dimensions["B"].width = 28
        worksheet.column_dimensions["C"].width = 24

    return csv_filename, excel_filename


@app.route("/", methods=["GET", "POST"])
def index():
    """Home page where users enter a company ticker symbol."""
    if request.method == "POST":
        ticker_symbol = request.form.get("ticker", "").upper().strip()
        if not ticker_symbol:
            return render_template("index.html", error="Please enter a ticker symbol.")

        try:
            data = get_financial_data(ticker_symbol)
            ratios = calculate_ratios(data)
            category = categorize_company(data, ratios)
            chart_file = make_chart(ticker_symbol, data, ratios)
            csv_file, excel_file = save_outputs(ticker_symbol, data, ratios, category)

            display_data = {key: format_currency(value) if key in [
                "Current Price", "Market Cap", "Total Revenue", "Gross Profits",
                "Net Income", "Total Cash", "Total Debt", "Book Value"
            ] else value for key, value in data.items()}

            display_ratios = {key: format_number(value) for key, value in ratios.items()}

            return render_template(
                "results.html",
                ticker=ticker_symbol,
                data=display_data,
                ratios=display_ratios,
                category=category,
                chart_file=chart_file,
                csv_file=csv_file,
                excel_file=excel_file,
            )
        except Exception as error:
            return render_template("index.html", error=str(error))

    return render_template("index.html")


@app.route("/download/<filename>")
def download_file(filename):
    """Lets users download the CSV or Excel file created by the app."""
    file_path = os.path.join(OUTPUT_FOLDER, filename)
    if not os.path.exists(file_path):
        return "File not found", 404
    return send_file(file_path, as_attachment=True)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
